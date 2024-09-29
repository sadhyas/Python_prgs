from openpyxl import Workbook, load_workbook
import time

# creating excel workbook

book= Workbook ()
sheet= book.active

sheet ['A1'] = 56

sheet ['A2'] = 43

now =time.strftime("%x")
sheet ['A3'] = now

book.save ("sample2.xlsx")

wb=load_workbook ("sample2.xlsx")

sheet=wb.worksheets [0]
sheet2=wb.create_sheet ('demo')

name=["John", "Dennis", "Ferb", "Candas", "Tadashi"]

salary=["$45000", "$78000", "$34000", "$71000", "$98000"]

sheet2.cell (row=1, column=1).value="Name"
sheet2.cell (row=1, column=2).value="Salary" 
j=2

for i in range (5):
    sheet2.cell (row=j, column=1).value=name[i]
    sheet2.cell (row=j, column=2).value=salary[i]
    j+=1

# 2nd way of writing to cells

print (wb.active)

sheet.cell (row=6, column=9).value="Kumari"

sheet ['F10']=100

print (wb.active)

print (wb. sheetnames)

wb.save("sample2.xlsx") 

print ("Reading Rowl   :", sheet ['A1'].value)

print ("Reading Row3   :", sheet ['A3'].value)


cells=sheet2['A1':'B6']

for c1, c2 in cells:
    print (c1. value,'\t\t',c2.value)
